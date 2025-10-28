"""
Excel and CSV parser for extracting structured data from spreadsheets.

Supports:
- Excel files (.xlsx, .xls, .xlsm)
- CSV files (.csv)
- TSV files (.tsv)

Implements IParser interface.
"""

import os
from typing import Dict, Any, List, Optional, Union
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from ..interfaces.parser import IParser


class ExcelParser(IParser):
    """
    Excel and CSV parser for spreadsheet data extraction.
    
    Features:
    - Multi-sheet support
    - Formula evaluation
    - Data type detection
    - Named range extraction
    - Metadata extraction (author, dates, etc.)
    - Merged cell handling
    """
    
    def __init__(
        self,
        read_formulas: bool = False,
        evaluate_formulas: bool = True,
        skip_empty_sheets: bool = True
    ):
        """
        Initialize Excel parser.
        
        Args:
            read_formulas: Whether to read formulas (not just values)
            evaluate_formulas: Whether to evaluate formulas to get values
            skip_empty_sheets: Skip sheets with no data
        """
        self.read_formulas = read_formulas
        self.evaluate_formulas = evaluate_formulas
        self.skip_empty_sheets = skip_empty_sheets
    
    def extract_fields(self, file_path: str) -> Dict[str, Any]:
        """
        Extract data from Excel/CSV file.
        
        Args:
            file_path: Path to Excel or CSV file
            
        Returns:
            Dictionary with extracted data and metadata
            {
                'sheets': [
                    {
                        'name': 'Sheet1',
                        'headers': ['col1', 'col2', ...],
                        'rows': [[val1, val2, ...], ...],
                        'row_count': 100,
                        'column_count': 5,
                        'data_types': ['string', 'number', ...],
                        'has_formulas': True,
                        'named_ranges': ['TotalRevenue', ...]
                    },
                    ...
                ],
                'sheet_count': 3,
                'metadata': {
                    'title': '...',
                    'author': '...',
                    'created': '...',
                    'modified': '...'
                },
                'warnings': []
            }
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext in ['.csv', '.tsv']:
            return self._extract_from_csv(file_path, file_ext)
        elif file_ext in ['.xlsx', '.xlsm', '.xls']:
            return self._extract_from_excel(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_ext}")
    
    def _extract_from_csv(self, file_path: str, file_ext: str) -> Dict[str, Any]:
        """Extract data from CSV/TSV file."""
        try:
            # Determine delimiter
            delimiter = '\t' if file_ext == '.tsv' else ','
            
            # Read CSV
            df = pd.read_csv(file_path, delimiter=delimiter)
            
            # Convert to sheet format
            sheet_data = self._dataframe_to_sheet(df, 'Data')
            
            return {
                'sheets': [sheet_data],
                'sheet_count': 1,
                'metadata': {
                    'file_type': file_ext,
                    'file_size': os.path.getsize(file_path)
                },
                'warnings': []
            }
            
        except Exception as e:
            raise ValueError(f"Failed to parse CSV: {str(e)}")
    
    def _extract_from_excel(self, file_path: str) -> Dict[str, Any]:
        """Extract data from Excel file."""
        try:
            # Load workbook for metadata and formulas
            wb = openpyxl.load_workbook(
                file_path,
                data_only=self.evaluate_formulas,
                read_only=False
            )
            
            # Extract metadata
            metadata = self._extract_metadata(wb)
            
            # Read all sheets with pandas
            df_dict = pd.read_excel(file_path, sheet_name=None, header=None)
            
            # Process each sheet
            sheets_data = []
            warnings = []
            
            for sheet_name, df in df_dict.items():
                # Skip empty sheets if configured
                if self.skip_empty_sheets and df.empty:
                    warnings.append(f"Skipped empty sheet: {sheet_name}")
                    continue
                
                # Get corresponding openpyxl sheet for formulas/ranges
                ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
                
                # Convert to sheet format
                sheet_data = self._dataframe_to_sheet(df, sheet_name)
                
                # Add formula information if available
                if ws:
                    sheet_data['has_formulas'] = self._has_formulas(ws)
                    sheet_data['named_ranges'] = self._get_named_ranges(wb, sheet_name)
                    sheet_data['merged_cells'] = len(ws.merged_cells.ranges)
                
                sheets_data.append(sheet_data)
            
            return {
                'sheets': sheets_data,
                'sheet_count': len(sheets_data),
                'metadata': metadata,
                'warnings': warnings
            }
            
        except Exception as e:
            raise ValueError(f"Failed to parse Excel: {str(e)}")
    
    def _dataframe_to_sheet(self, df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """Convert pandas DataFrame to sheet dictionary."""
        if df.empty:
            return {
                'name': sheet_name,
                'headers': [],
                'rows': [],
                'row_count': 0,
                'column_count': 0,
                'data_types': [],
                'has_formulas': False,
                'named_ranges': []
            }
        
        # Detect if first row is headers
        first_row = df.iloc[0]
        is_header_row = self._is_header_row(first_row, df)
        
        if is_header_row:
            headers = [str(val) for val in first_row]
            data_df = df.iloc[1:]
        else:
            headers = [f"Column_{i+1}" for i in range(len(df.columns))]
            data_df = df
        
        # Convert to rows
        rows = data_df.values.tolist()
        
        # Clean values (convert NaN to empty string)
        rows = [
            ['' if pd.isna(val) else str(val) for val in row]
            for row in rows
        ]
        
        # Filter out completely empty rows
        rows = [row for row in rows if any(cell for cell in row)]
        
        # Detect data types
        data_types = self._detect_data_types(data_df)
        
        return {
            'name': sheet_name,
            'headers': headers,
            'rows': rows,
            'row_count': len(rows),
            'column_count': len(headers),
            'data_types': data_types,
            'has_formulas': False,  # Will be updated if openpyxl available
            'named_ranges': []
        }
    
    def _is_header_row(self, first_row: pd.Series, df: pd.DataFrame) -> bool:
        """
        Detect if first row is likely headers.
        
        Heuristics:
        - All string values
        - No duplicate values
        - Different from subsequent rows
        """
        # Check if all values are strings
        if not all(isinstance(val, str) or pd.isna(val) for val in first_row):
            # Has numbers, likely not headers
            pass
        
        # Check for duplicates
        non_null = [val for val in first_row if not pd.isna(val)]
        if len(non_null) != len(set(non_null)):
            return False  # Has duplicates
        
        # If more than 50% are strings and rest is numeric data, likely headers
        if len(df) > 1:
            first_row_strings = sum(1 for val in first_row if isinstance(val, str))
            second_row_strings = sum(1 for val in df.iloc[1] if isinstance(val, str))
            
            if first_row_strings > second_row_strings:
                return True
        
        # Default: assume first row is headers
        return True
    
    def _detect_data_types(self, df: pd.DataFrame) -> List[str]:
        """Detect data types for each column."""
        data_types = []
        
        for col in df.columns:
            col_data = df[col].dropna()
            
            if col_data.empty:
                data_types.append('empty')
                continue
            
            # Check pandas dtype
            dtype = col_data.dtype
            
            if pd.api.types.is_integer_dtype(dtype):
                data_types.append('integer')
            elif pd.api.types.is_float_dtype(dtype):
                data_types.append('float')
            elif pd.api.types.is_bool_dtype(dtype):
                data_types.append('boolean')
            elif pd.api.types.is_datetime64_any_dtype(dtype):
                data_types.append('datetime')
            elif pd.api.types.is_string_dtype(dtype) or pd.api.types.is_object_dtype(dtype):
                # Check if it's currency
                sample = str(col_data.iloc[0])
                if '$' in sample or '€' in sample or '£' in sample:
                    data_types.append('currency')
                # Check if it's percentage
                elif '%' in sample:
                    data_types.append('percentage')
                else:
                    data_types.append('string')
            else:
                data_types.append('unknown')
        
        return data_types
    
    def _extract_metadata(self, wb: openpyxl.Workbook) -> Dict[str, Any]:
        """Extract workbook metadata."""
        props = wb.properties
        
        metadata = {
            'title': props.title or '',
            'author': props.creator or '',
            'subject': props.subject or '',
            'keywords': props.keywords or '',
            'description': props.description or '',
            'category': props.category or '',
            'comments': props.comments or ''
        }
        
        if props.created:
            metadata['created'] = props.created.isoformat()
        if props.modified:
            metadata['modified'] = props.modified.isoformat()
        
        metadata['sheet_names'] = wb.sheetnames
        
        return metadata
    
    def _has_formulas(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> bool:
        """Check if worksheet has any formulas."""
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Formula
                    return True
        return False
    
    def _get_named_ranges(
        self,
        wb: openpyxl.Workbook,
        sheet_name: str
    ) -> List[str]:
        """Get named ranges for a specific sheet."""
        named_ranges = []
        
        for name in wb.defined_names.definedName:
            # Check if range belongs to this sheet
            if sheet_name in str(name.value):
                named_ranges.append(name.name)
        
        return named_ranges
    
    def is_fillable(self, file_path: str) -> bool:
        """
        Excel parser doesn't work with fillable PDF fields.
        
        Args:
            file_path: Path to file
            
        Returns:
            Always False
        """
        return False
    
    def get_sheet_by_name(
        self,
        result: Dict[str, Any],
        sheet_name: str
    ) -> Optional[Dict[str, Any]]:
        """
        Get sheet data by name.
        
        Args:
            result: Extraction result from extract_fields()
            sheet_name: Name of sheet to find
            
        Returns:
            Sheet data or None if not found
        """
        for sheet in result['sheets']:
            if sheet['name'].lower() == sheet_name.lower():
                return sheet
        return None
    
    def find_column_by_header(
        self,
        sheet: Dict[str, Any],
        header_keyword: str
    ) -> Optional[int]:
        """
        Find column index by header keyword.
        
        Args:
            sheet: Sheet data dictionary
            header_keyword: Keyword to search for in headers (case-insensitive)
            
        Returns:
            Column index (0-based) or None if not found
        """
        headers_lower = [h.lower() for h in sheet['headers']]
        
        for idx, header in enumerate(headers_lower):
            if header_keyword.lower() in header:
                return idx
        
        return None
    
    def get_column_values(
        self,
        sheet: Dict[str, Any],
        column_index: int,
        skip_empty: bool = True
    ) -> List[str]:
        """
        Get all values from a specific column.
        
        Args:
            sheet: Sheet data dictionary
            column_index: Column index (0-based)
            skip_empty: Skip empty values
            
        Returns:
            List of column values
        """
        values = []
        
        for row in sheet['rows']:
            if column_index < len(row):
                value = row[column_index]
                if skip_empty and not value:
                    continue
                values.append(value)
        
        return values
    
    def __repr__(self) -> str:
        """String representation."""
        return (
            f"ExcelParser(read_formulas={self.read_formulas}, "
            f"evaluate_formulas={self.evaluate_formulas})"
        )